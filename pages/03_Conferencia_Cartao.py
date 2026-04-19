import streamlit as st
import pandas as pd
import pdfplumber
from io import BytesIO
import openpyxl
from datetime import timedelta
import re

st.set_page_config(page_title="Conferência Cielo Total", layout="wide")

st.title("💳 Gestão Financeira - Conferência Cielo 20/20")
st.markdown("---")

u_excel = st.file_uploader("1. Selecione a Planilha Cielo", type=['xlsx'])
u_pdf = st.file_uploader("2. Selecione o Relatório do Sistema (PDF)", type=['pdf'])

if u_excel and u_pdf:
    # --- MAPEAMENTO INICIAL DO PDF ---
    @st.cache_data
    def extrair_dados_pdf(pdf_file):
        base = []
        with pdfplumber.open(pdf_file) as pdf:
            for page in pdf.pages:
                texto = page.extract_text()
                if texto:
                    for linha in texto.split('\n'):
                        # Captura ID com símbolos (ex: PC22630-1)
                        match_id = re.search(r'(PC|PD)[^\s]+', linha, re.IGNORECASE)
                        match_data = re.search(r'\d{2}/\d{2}/\d{4}', linha)
                        
                        if match_id:
                            cod_id = match_id.group().strip().upper()
                            dt_pdf = pd.to_datetime(match_data.group(), dayfirst=True).date() if match_data else None
                            
                            # Captura valores (ex: 156 ou 156,00)
                            numeros = re.findall(r'\d+(?:[\.,]\d{2})?|\d+', linha)
                            for n in numeros:
                                try:
                                    v_f = float(n.replace('.', '').replace(',', '.'))
                                    if v_f > 1.0:
                                        base.append({'id': cod_id, 'valor': v_f, 'data': dt_pdf, 'usado': False})
                                except: continue
        return base

    lista_pdf = extrair_dados_pdf(u_pdf)

    # Criamos um estado para a planilha para que os botões trabalhem no mesmo objeto
    if 'wb' not in st.session_state:
        u_excel.seek(0)
        st.session_state.wb = openpyxl.load_workbook(u_excel)
        st.session_state.processado = False

    wb = st.session_state.wb
    ws = wb.active
    df_cielo = pd.read_excel(u_excel, header=14)

    col1, col2 = st.columns(2)

    # --- BOTÃO 1: BUSCA PADRÃO ---
    if col1.button("🚀 1. Executar Busca Padrão"):
        sucessos = 0
        for i, row in df_cielo.iterrows():
            linha_ex = i + 16
            val_cielo = float(row.iloc[4]) # Valor Bruto
            dt_venda = pd.to_datetime(row.iloc[1]).date()
            
            for item in lista_pdf:
                if not item['usado'] and abs(item['valor'] - val_cielo) <= 0.01:
                    if item['data'] and abs((item['data'] - dt_venda).days) <= 1:
                        ws.cell(row=linha_ex, column=8).value = item['id']
                        item['usado'] = True
                        sucessos += 1
                        break
        st.session_state.processado = True
        st.success(f"Busca padrão concluída! Itens achados: {sucessos}")

    # --- BOTÃO 2: BUSCA AVANÇADA (SÓ NO QUE FALTOU) ---
    if col2.button("🔍 2. Forçar Busca nos Restantes"):
        recuperados = 0
        for i, row in df_cielo.iterrows():
            linha_ex = i + 16
            # Só mexe se estiver vazio ou com erro da busca anterior
            status_atual = ws.cell(row=linha_ex, column=8).value
            if status_atual in [None, "NÃO ENCONTRADO", "REVISAR"]:
                val_cielo = float(row.iloc[4])
                dt_venda = pd.to_datetime(row.iloc[1]).date()
                
                for item in lista_pdf:
                    if not item['usado']:
                        # Critérios extras: data até 3 dias e valor até 0.03
                        v_bate = abs(item['valor'] - val_cielo) <= 0.03
                        d_bate = True
                        if item['data']:
                            d_bate = abs((item['data'] - dt_venda).days) <= 3
                        
                        if v_bate and d_bate:
                            ws.cell(row=linha_ex, column=8).value = item['id']
                            item['usado'] = True
                            recuperados += 1
                            break
                
                # Se ainda assim não achar, marca para revisão
                if ws.cell(row=linha_ex, column=8).value is None:
                    ws.cell(row=linha_ex, column=8).value = "NÃO ENCONTRADO"
                    
        st.success(f"Busca avançada concluída! Itens recuperados: {recuperados}")

    # --- DOWNLOAD ÚNICO ---
    st.markdown("---")
    buffer = BytesIO()
    wb.save(buffer)
    st.download_button(
        label="📥 Baixar Planilha Final (20/20)",
        data=buffer.getvalue(),
        file_name="Conferencia_Cielo_Consolidada.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

    if st.button("♻️ Reiniciar Processo"):
        for key in st.session_state.keys():
            del st.session_state[key]
        st.rerun()
